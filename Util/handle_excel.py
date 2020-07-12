import openpyxl
import os

base_path=os.path.dirname(os.getcwd())

class HandExcel:
    def load_excel(self):
        '''
        加载excel
        :return:
        '''
        open_excel = openpyxl.load_workbook(base_path+r'\Case\imooc.xlsx')
        return open_excel


    def get_sheet_data(self,index=None):
        '''
        加载所有sheet内容
        :return:
        '''
        sheet_name =self.load_excel().sheetnames
        if index == None:
            index = 0
        data = self.load_excel()[sheet_name[index]]
        return data


    def get_cell_value(self,row,col):
        '''
        获取单元格里的内容
        :param row:
        :param col:
        :return:
        '''
        data = self.get_sheet_data().cell(row=row,column=col)

        return data.value

    def get_rows(self):
        '''
        获取行数
        :return:
        '''
        row = self.get_sheet_data().max_row
        return row

    def get_rows_value(self,row):
        '''
        获取某一行的内容
        :param row:
        :return:
        '''

        row_list = []
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list

    def excel_write_data(self,rows,cols,value):
        '''
        写入数据
        :param rows:
        :param cols:
        :param value:
        :return:
        '''
        wb = self.load_excel()
        wr = wb.active
        wr.cell(rows,cols,value)
        wb.save(base_path+r'\Case\imooc.xlsx')

    def get_column_value(self,key=None):
        '''
        获取某一列
        :param key:
        :return:
        '''
        if key ==None:
            key = 'A'
        column_list = []
        column_list_data = self.get_sheet_data()[key]
        for i in column_list_data:
            column_list.append(i.value)
        return column_list

    def get_rows_number(self,case_id):
        '''
        获取行号
        :return:
        '''
        number=1
        cols_data = self.get_column_value()
        for col_data in  cols_data:
            if case_id == col_data:
                return number
            number = number+1
        return number

    def get_excel_data(self):
        '''
        获取excel里的所有的数据
        :return:
        '''
        data_list= []
        for i in range(self.get_rows()):
            data_list.append(self.get_rows_value(i+2))

        return data_list




excel_data = HandExcel()

if __name__ == '__main__':
    handle = HandExcel()
    #print(handle.get_cell_value(1,3).value)
    #print(handle.get_rows_value(2))
    #handle.excel_write_data(2,11,'通过')
    #print(handle.get_column_value('A'))
    #print(handle.get_rows_number('imooc_003'))
    print(handle.get_excel_data())